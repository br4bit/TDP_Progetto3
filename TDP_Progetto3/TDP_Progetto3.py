import TdP_collections.hash_table.sorted_table_map
from builtins import print
from openpyxl.reader.excel import load_workbook
from builtins import list
from builtins import range
from openpyxl import cell
import openpyxl
from openpyxl import workbook

def fail_func(w):
    #initializing the array auxiliar with 0 in each cell
    auxiliar = [0] * len(w)
    # for index 0, it will always be 0, starting from 1
    i = 1
    #m can also be viewed as index of first mismatch
    m = 0
    while i < len(w):
    # prefix = suffix till m-1
        if w[i] == w[m]:
            m += 1
            auxiliar[i] = m
            i += 1
            # when there is a mismatch,we will check the index of previous possible prefix.
        elif w[i] != w[m] and m != 0:
        # Note that we do not increment i here.
            m = auxiliar[m-1]
        else:
                # m = 0, we move to the next letter, 
                # there was no any prefix found which is equal to the suffix for index i
            auxiliar[i] = 0
            i += 1
    return auxiliar

def count_kmp(T,P):
    count = 0
    array_aux = fail_func(P)
    #counter for word P
    i = 0
    #counter for word T
    j = 0
    while j < len(T):
        #we need handle 2 conditions when there is a mismatch
        if P[i] != T[j]:
            #1s condition
            if i == 0:
                #start again from next character in the text T
                j += 1
            else:
                # array_aux[i-1] will tell from where to compare next
                # and no need to match for P[0..array_aux[i-1] - 1],
                # they will match anyway.
                i = array_aux[i-1]
        else:
            i += 1
            j += 1
            # we found the pattern
            if i == len(P):
                #count + 1
                count+=1
                i = array_aux[i-1]
    return count

def load_excel():
    print("Loading xls...")
    wb = load_workbook('all-euro-data-2016-2017.xlsx')
    print("Loading done...")
    lista=list()
    sheet = wb.get_sheet_by_name('E0')
    for i in range(1, sheet.max_column):
        value=sheet.cell(row=1,column=i).value
        if(value == 'HomeTeam'):
             for i in range(2, sheet.max_row):
                b=sheet.cell(row=i, column=3).value
                if b not in lista:
                    lista.append(b)
             break
        else:
            continue     
    print(lista)
    print(sheet.max_row)
    # Print out values in column 2 
    #for i in range(1, 11):
        #print(i, sheet.cell(row=1, column=i).value)

    

def main():
    pattern = "acabacacd"
    text = "acfacabacabacacdkacabacacdacabacacd"
    print("Text: ",text)
    print("Pattern: ",pattern)
    print("Numero di occorrenze KMP algorithm: ",count_kmp(text,pattern))
    print()
    load_excel()

if __name__ == '__main__':
    main()
    